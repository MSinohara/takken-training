import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / ".tmp" / "development-sheet.xlsx"
DESTINATION = ROOT / ".tmp" / "development-data.json"
SHEETS = ("会員マスタ", "個人会員マスタ", "研修会", "受付索引")


def read_records(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows)]
    records = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        records.append({header: value for header, value in zip(headers, row) if header})
    return records


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
data = {sheet: read_records(workbook[sheet]) for sheet in SHEETS}
DESTINATION.write_text(json.dumps(data, ensure_ascii=False, default=str), encoding="utf-8")
print(" / ".join(f"{name}: {len(rows)}件" for name, rows in data.items()))
